#!/usr/bin/env python3
"""
generate_excel.py — 从 data/professors/*.json 生成 comparison.xlsx
支持多 Sheet：对比总表 / 论文列表 / 学生信息 / 课题组
"""

import json
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "professors")
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "docs")

# Color definitions
HEADER_FILL = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HIGH_QUALITY_FILL = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
MEDIUM_QUALITY_FILL = PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid")
LOW_QUALITY_FILL = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
RECRUITING_FILL = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="top")


def load_professors():
    """Load all professor JSON files from data/professors/"""
    professors = []
    index_path = os.path.join(DATA_DIR, "index.json")
    if not os.path.exists(index_path):
        print(f"Warning: {index_path} not found")
        return professors

    with open(index_path, "r", encoding="utf-8") as f:
        index = json.load(f)

    for entry in index:
        prof_path = os.path.join(DATA_DIR, f"{entry['id']}.json")
        if os.path.exists(prof_path):
            with open(prof_path, "r", encoding="utf-8") as f:
                professors.append(json.load(f))
    return professors


def safe_str(value, lang="zh"):
    """Safely extract string value, handling bilingual dicts"""
    if value is None:
        return ""
    if isinstance(value, dict):
        result = value.get(lang) or value.get("en") or ""
        return result if result else ""
    if isinstance(value, list):
        if len(value) > 0 and isinstance(value[0], dict):
            return ", ".join(safe_str(v, lang) for v in value)
        return ", ".join(str(v) for v in value if v is not None)
    return str(value) if value else ""


def apply_header_style(ws, row, max_col):
    """Apply header styling to a row"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER


def apply_data_style(ws, start_row, end_row, max_col):
    """Apply data cell styling"""
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT


def generate_comparison_sheet(wb, professors):
    """Sheet 1: 对比总表"""
    ws = wb.active
    ws.title = "对比总表"

    headers = [
        "姓名", "英文名", "职称", "院校", "学院", "地区",
        "邮箱", "电话", "研究方向(主页)", "研究方向(论文推断)",
        "课题组", "是否招生", "招生要求",
        "近三年论文数", "h-index", "学生数(推断)",
        "数据质量", "最后更新"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for i, prof in enumerate(professors):
        row = i + 2
        r = prof.get("research", {})
        c = prof.get("contact", {})
        g = prof.get("group", {})
        a = prof.get("admission", {})
        p = prof.get("publications", {})
        s = prof.get("students", {})
        q = prof.get("_quality", {})

        values = [
            safe_str(prof.get("name", {})),
            safe_str(prof.get("name", {}), "en"),
            prof.get("title", ""),
            safe_str(prof.get("university", {})),
            safe_str(prof.get("department", {})),
            prof.get("region", ""),
            c.get("email", ""),
            c.get("phone", ""),
            ", ".join(safe_str(d) for d in r.get("homepage_directions", [])),
            ", ".join(r.get("inferred_from_papers", [])),
            safe_str(g.get("name", {})),
            "是" if a.get("is_recruiting") else ("否" if a.get("is_recruiting") is False else "未知"),
            safe_str(a.get("requirements", {})).replace("\n", " ")[:200],
            len(p.get("recent", [])),
            p.get("h_index", ""),
            len(s.get("inferred_from_papers", [])),
            q.get("overall", "unknown"),
            prof.get("_sources", {}).get("fetched_at", "")
        ]

        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col, value=value)

    apply_header_style(ws, 1, len(headers))
    apply_data_style(ws, 2, len(professors) + 1, len(headers))

    # Auto-fit column widths (approximate)
    for col in range(1, len(headers) + 1):
        max_width = len(str(ws.cell(row=1, column=col).value)) * 2
        for row in range(2, min(len(professors) + 2, 50)):
            cell_val = str(ws.cell(row=row, column=col).value)
            max_width = max(max_width, min(len(cell_val) * 1.2, 60))
        ws.column_dimensions[get_column_letter(col)].width = max_width + 2

    # Freeze header row & add auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(professors) + 1}"

    # Conditional formatting for data quality
    quality_col = get_column_letter(headers.index("数据质量") + 1)
    ws.conditional_formatting.add(
        f"{quality_col}2:{quality_col}{len(professors) + 1}",
        CellIsRule(operator="equal", formula=['"high"'], fill=HIGH_QUALITY_FILL)
    )
    ws.conditional_formatting.add(
        f"{quality_col}2:{quality_col}{len(professors) + 1}",
        CellIsRule(operator="equal", formula=['"medium"'], fill=MEDIUM_QUALITY_FILL)
    )
    ws.conditional_formatting.add(
        f"{quality_col}2:{quality_col}{len(professors) + 1}",
        CellIsRule(operator="equal", formula=['"low"'], fill=LOW_QUALITY_FILL)
    )

    # Conditional formatting for recruiting status
    recruit_col = get_column_letter(headers.index("是否招生") + 1)
    ws.conditional_formatting.add(
        f"{recruit_col}2:{recruit_col}{len(professors) + 1}",
        CellIsRule(operator="equal", formula=['"是"'], fill=RECRUITING_FILL)
    )


def generate_publications_sheet(wb, professors):
    """Sheet 2: 论文列表"""
    ws = wb.create_sheet("论文列表")

    headers = ["导师", "论文标题", "作者", "期刊/会议", "年份", "类型", "引用数", "URL"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    row = 2
    for prof in professors:
        prof_name = safe_str(prof.get("name", {}))
        publications = prof.get("publications", {}).get("recent", [])
        for pub in publications:
            authors = pub.get("authors", [])
            if isinstance(authors, list):
                author_str = ", ".join(
                    a.get("name", str(a)) if isinstance(a, dict) else str(a)
                    for a in authors
                )
            else:
                author_str = str(authors)

            values = [
                prof_name,
                pub.get("title", ""),
                author_str,
                pub.get("venue", ""),
                pub.get("year", ""),
                pub.get("type", ""),
                pub.get("citations", 0),
                pub.get("url", "")
            ]
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1

    if row > 2:
        apply_header_style(ws, 1, len(headers))
        apply_data_style(ws, 2, row - 1, len(headers))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25


def generate_students_sheet(wb, professors):
    """Sheet 3: 学生信息"""
    ws = wb.create_sheet("学生信息")

    headers = ["导师", "学生姓名", "推断来源", "作为一作论文数", "一作论文列表"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    row = 2
    for prof in professors:
        prof_name = safe_str(prof.get("name", {}))
        students = prof.get("students", {}).get("inferred_from_papers", [])
        for stu in students:
            if isinstance(stu, dict):
                values = [
                    prof_name,
                    stu.get("name", ""),
                    "论文一作反推",
                    stu.get("papers_count", ""),
                    ", ".join(stu.get("first_author_in", []))
                ]
            else:
                values = [prof_name, str(stu), "", "", ""]
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1

    if row > 2:
        apply_header_style(ws, 1, len(headers))
        apply_data_style(ws, 2, row - 1, len(headers))
        ws.freeze_panes = "A2"

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30


def generate_groups_sheet(wb, professors):
    """Sheet 4: 课题组"""
    ws = wb.create_sheet("课题组")

    headers = ["课题组名称", "成员姓名", "院校", "来源", "课题组URL"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    row = 2
    for prof in professors:
        group = prof.get("group", {})
        group_name = safe_str(group.get("name", {}))
        if group_name:
            values = [
                group_name,
                safe_str(prof.get("name", {})),
                safe_str(prof.get("university", {})),
                group.get("source", ""),
                group.get("url", "")
            ]
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1

    if row > 2:
        apply_header_style(ws, 1, len(headers))
        apply_data_style(ws, 2, row - 1, len(headers))
        ws.freeze_panes = "A2"

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30


def generate_excel(professors, output_path):
    """Main function: generate the Excel workbook"""
    wb = Workbook()
    generate_comparison_sheet(wb, professors)
    generate_publications_sheet(wb, professors)
    generate_students_sheet(wb, professors)
    generate_groups_sheet(wb, professors)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Excel generated: {output_path}")
    return output_path


def main():
    professors = load_professors()
    if not professors:
        print("No professor data found. Please add professors first.")
        sys.exit(1)

    output_path = os.path.join(OUTPUT_DIR, "comparison.xlsx")
    generate_excel(professors, output_path)
    print(f"Done. {len(professors)} professors exported to Excel.")


if __name__ == "__main__":
    main()
